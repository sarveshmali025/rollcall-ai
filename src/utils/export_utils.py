import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def _rows_from_logs(logs):
    """Normalizes raw attendance_logs rows (with joined students(name)) into
    a flat list of (name, student_id, timestamp, status) tuples, sorted by
    timestamp then name."""
    rows = []
    for log in logs:
        student_info = log.get('students') or {}
        rows.append((
            student_info.get('name', f"Student {log.get('student_id')}"),
            log.get('student_id'),
            log.get('timestamp'),
            'Present' if log.get('is_present') else 'Absent',
        ))
    rows.sort(key=lambda r: (r[2] or '', r[0] or ''))
    return rows


def build_excel_bytes(subject_name, logs):
    """Returns raw xlsx bytes for the given subject's attendance logs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Student Name", "Student ID", "Session Timestamp", "Status"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    rows = _rows_from_logs(logs)
    for row in rows:
        ws.append(row)

    present_fill = PatternFill(start_color="D7F7E3", end_color="D7F7E3", fill_type="solid")
    absent_fill = PatternFill(start_color="FCE0E1", end_color="FCE0E1", fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row_idx, column=4)
        status_cell.fill = present_fill if status_cell.value == 'Present' else absent_fill

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)])
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"

    summary_ws = wb.create_sheet("Summary")
    total = len(rows)
    present = sum(1 for r in rows if r[3] == 'Present')
    pct = round(100 * present / total, 1) if total else 0.0

    summary_ws.append(["Subject", subject_name])
    summary_ws.append(["Total Records", total])
    summary_ws.append(["Present", present])
    summary_ws.append(["Absent", total - present])
    summary_ws.append(["Attendance %", f"{pct}%"])
    summary_ws.append(["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_pdf_bytes(subject_name, logs):
    """Returns raw PDF bytes for the given subject's attendance logs."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=16 * mm, rightMargin=16 * mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'RCTitle', parent=styles['Title'], textColor=colors.HexColor('#5865F2'), fontSize=20
    )
    meta_style = ParagraphStyle(
        'RCMeta', parent=styles['Normal'], textColor=colors.HexColor('#64678C'), fontSize=9
    )

    elements = [
        Paragraph("RollCall AI &mdash; Attendance Report", title_style),
        Spacer(1, 4),
        Paragraph(f"Subject: {subject_name}", styles['Heading3']),
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style),
        Spacer(1, 12),
    ]

    rows = _rows_from_logs(logs)
    total = len(rows)
    present = sum(1 for r in rows if r[3] == 'Present')
    pct = round(100 * present / total, 1) if total else 0.0

    summary_data = [
        ["Total Records", "Present", "Absent", "Attendance %"],
        [str(total), str(present), str(total - present), f"{pct}%"],
    ]
    summary_table = Table(summary_data, hAlign='LEFT', colWidths=[35 * mm] * 4)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EB459E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E3FF')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 16))

    table_data = [["Student Name", "Student ID", "Session Timestamp", "Status"]]
    table_data += [[str(c) for c in row] for row in rows]

    detail_table = Table(table_data, hAlign='LEFT', repeatRows=1, colWidths=[55 * mm, 25 * mm, 55 * mm, 25 * mm])
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5865F2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E0E3FF')),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]
    for row_idx, row in enumerate(rows, start=1):
        color = colors.HexColor('#D7F7E3') if row[3] == 'Present' else colors.HexColor('#FCE0E1')
        style_cmds.append(('BACKGROUND', (3, row_idx), (3, row_idx), color))

    detail_table.setStyle(TableStyle(style_cmds))
    elements.append(detail_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
